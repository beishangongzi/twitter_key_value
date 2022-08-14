import pandas as pd
from openpyxl import load_workbook


# 将dict类型的数据追加写入到现有的Excel中
def write_to_exist_excel1(fileName, sheetName):
    df_old = pd.DataFrame(pd.read_excel(fileName, sheet_name=sheetName))  # 读取原数据文件和表
    row_old = df_old.shape[0]  # 获取原数据的行数

    data_added = {"a": 99, "b": 98, "c": 97}
    df = pd.DataFrame(data_added, index=[0])  # 如果dict的value只有一行，加上index = [0]即可

    book = load_workbook(fileName)
    writer = pd.ExcelWriter(fileName, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # 将data_added数据写入Excel中
    df.to_excel(writer, sheet_name=sheetName, startrow=row_old + 1, index=False, header=False)

    writer.save()  # 保存


# 将list[dict]类型的数据追加写入到现有的Excel中
def write_to_exist_excel2(fileName, sheetName, data_added):
    df_old = pd.DataFrame(pd.read_excel(fileName, sheet_name=sheetName))  # 读取原数据文件和表
    row_old = df_old.shape[0]  # 获取原数据的行数

    # data_added = [{"a": 9, "b": 8, "c": 7, "d": 6},
    #               {"a": 37, "b": 38, "c": 39, "d": 40}]
    df = pd.DataFrame(data_added)

    book = load_workbook(fileName)
    writer = pd.ExcelWriter(fileName, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # 将data_added数据写入Excel中
    df.to_excel(writer, sheet_name=sheetName, startrow=row_old + 1, index=False, header=False)

    writer.save()  # 保存


if __name__ == '__main__':
    fileName = 'data.xlsx'
    sheetName = 'Sheet1'
    write_to_exist_excel1(fileName, sheetName)
    write_to_exist_excel2(fileName, sheetName)